from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


class XlPainter():

    col_lables = [chr(i) for i in range(65, 91)]
    col_lables = col_lables + [ chr(a)+chr(b) for a in range(65, 91) for b in range(65,91)]

    redFill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type = 'solid')
    grnFill = PatternFill(start_color='00ee00', end_color='00ee00', fill_type='solid' )
    whiteFill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
    yellowFill = PatternFill(start_color='FFDF00', end_color='FFDF00', fill_type='solid')

    def __init__(self, outfile=None):
        if outfile is None:
            self.wb = Workbook()
            self.new_state = True
        else:
            self.wb = load_workbook(outfile)
            self.new_state = False
        self.ws = self.wb.active
        self.outfile = outfile

        self.alignment = Alignment(wrap_text=True)
        self.header_color = 'b7dee8'
        self.header_font = Font(bold=True, size=12)

    def write_DF(self, df, col_widths=None, wrap_columns=None, hide_columns=None, index=False, header=False,
                 outfile=None, sheet_name='Sheet1', cf=None):

        if not self.new_state:
            self.ws = self.wb.create_sheet(sheet_name)
            self.new_state = False

        self.ws.title = sheet_name

        if col_widths is not None:
            self.__set_col_widths(col_widths)

        for r in dataframe_to_rows(df, index=index, header=header):
            self.ws.append(r)

        for i in range(len(df.columns)):
            cell = self.ws[f'{self.col_lables[i]}1']
            cell.fill = PatternFill(start_color=self.header_color, fill_type = "solid")
            cell.font = self.header_font

        if wrap_columns is not None:
            for col in wrap_columns:
                for i in range(len(df.index)+1): # add 1 for header row
                    cell = self.ws[f'{col}{i + 1}'] # add 1 to make it 1-index
                    cell.alignment = self.alignment

        if hide_columns is not None:
            for col in hide_columns:
                self.ws.column_dimensions[col].hidden = True

        if cf is not None:
            self.__apply_cf(cf, len(df.index))

        if outfile is None: outfile = self.outfile
        if outfile is not None:
            self.wb.save(outfile)

    def __set_col_widths(self, col_widths):
        for i, width in enumerate(col_widths):
            col = self.ws.column_dimensions[self.col_lables[i]]
            col.width = width
            col.alignment = self.alignment

    def __apply_cf(self, col_fmts: dict, nrows):
        cols = col_fmts.keys()
        for key, val in col_fmts.items():
            for formula, pattern in val:
                start = f'{key}{2}' #1 left for header row
                end = f'{key}{nrows+1}'
                self.ws.conditional_formatting.add(f'{start}:{end}',FormulaRule(formula=[f'{start}{formula}'], fill=pattern))
